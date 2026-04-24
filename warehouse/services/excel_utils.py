"""
Утиліти для генерації Excel-файлів.

Єдине місце для стилів, захисту від formula injection та формування відповіді.
Використовується у views/reports.py, views/orders.py та аналітичних вьюхах.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from decimal import Decimal


_THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

_CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def sanitize_cell(value):
    """Запобігає Excel formula injection (= + - @ на початку рядка)."""
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value
    return value


def create_excel_response(
    headers,
    data_rows,
    filename,
    sheet_title="Report",
    header_color="4F81BD",
    autosize=True,
    fixed_col_width=None,
):
    """
    Створює Excel-файл зі стилізованими заголовками та повертає HttpResponse.

    Args:
        headers:        список рядків — заголовки колонок.
        data_rows:      ітерабельне, де кожен елемент — список/кортеж значень рядка.
        filename:       ім'я файлу для Content-Disposition (без лапок, без пробілів).
        sheet_title:    назва листа.
        header_color:   hex-рядок кольору фону заголовка (без '#'), default '4F81BD'.
        autosize:       якщо True — автоматично підганяє ширину колонок за вмістом.
        fixed_col_width:якщо задано (int), всі колонки матимуть цю ширину (autosize ігнорується).

    Returns:
        HttpResponse з application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")

    # Рядок заголовків
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER_ALIGN

    # Рядки даних
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_cell(value))
            cell.border = _THIN_BORDER
            if isinstance(value, (int, float, Decimal)):
                cell.alignment = Alignment(horizontal='right')

    # Ширина колонок
    if fixed_col_width is not None:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = fixed_col_width
    elif autosize:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
