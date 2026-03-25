from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json
from decimal import Decimal
from ..models import StageLimit, Transaction

@login_required
def mechanisms_analytics(request):
    """
    Звіт по Механізмах (Спецтехніка).
    Виправлено: AttributeError 'StageLimit' object has no attribute 'construct_type'.
    Тепер тип визначається через категорію матеріалу.
    """
    # 1. Шукаємо ліміти по категорії "техніка" (або схоже)
    # Оптимізація: додано material__category до select_related
    limits = StageLimit.objects.filter(
        material__category__name__icontains='техніка'
    ).select_related('stage', 'material', 'stage__warehouse', 'material__category').order_by('stage__name')
    
    report_data = []
    total_plan = Decimal("0.000")
    total_fact = Decimal("0.000")
    
    # Масиви для графіка Chart.js
    chart_labels = []
    chart_plan_data = []
    chart_fact_data = []
    
    for limit in limits:
        # Рахуємо фактичне відпрацювання
        fact_spent = Transaction.objects.filter(
            warehouse=limit.stage.warehouse,
            material=limit.material,
            transaction_type='OUT',
            stage=limit.stage
        ).aggregate(s=Sum('quantity'))['s'] or Decimal("0.000")
        
        plan = limit.planned_quantity
        fact = fact_spent
        diff = plan - fact
        
        status = 'ok'
        if diff < 0: 
            status = 'over'
        elif plan > 0 and diff < (plan * Decimal("0.1")): 
            status = 'warning'
            
        percent = (fact / plan * 100) if plan > 0 else 0
        
        # ВИПРАВЛЕННЯ: Визначаємо тип механізму через категорію матеріалу
        type_name = "Без категорії"
        if limit.material.category:
            type_name = limit.material.category.name
        
        report_data.append({
            'warehouse': limit.stage.warehouse.name,
            'stage': limit.stage.name,
            'material': limit.material.name,
            'characteristics': limit.material.characteristics,
            'unit': limit.material.unit,
            'plan': plan,
            'fact': fact,
            'diff': diff,
            'percent': int(percent),
            'status': status,
            'type': type_name  # Використовуємо category.name замість construct_type
        })
        
        total_plan += plan
        total_fact += fact
        
        # Data for chart
        chart_labels.append(f"{limit.stage.name} ({limit.material.name})")
        chart_plan_data.append(float(plan))
        chart_fact_data.append(float(fact))

    # EXPORT TO EXCEL
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Звіт по механізмах"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="E07A5F", end_color="E07A5F", fill_type="solid")
        center_align = Alignment(horizontal='center')
        
        headers = ['Об\'єкт', 'Етап', 'Механізм', 'Характеристики', 'Од.', 'План', 'Факт', 'Різниця', 'Статус']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row in report_data:
            st = "Норма"
            if row['status'] == 'over': st = "ПЕРЕПРАЦЮВАННЯ"
            elif row['status'] == 'warning': st = "Увага"
            
            ws.append([
                row['warehouse'], row['stage'], row['material'], 
                row['characteristics'], row['unit'], row['plan'], 
                row['fact'], row['diff'], st
            ])
            
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Mechanisms_Report.xlsx'
        wb.save(response)
        return response

    return render(request, 'warehouse/mechanisms_report.html', {
        'report_data': report_data,
        'total_plan': total_plan,
        'total_fact': total_fact,
        'total_diff': total_plan - total_fact,
        # JSON для графіка
        'chart_labels': json.dumps(chart_labels),
        'chart_plan_data': json.dumps(chart_plan_data),
        'chart_fact_data': json.dumps(chart_fact_data),
    })