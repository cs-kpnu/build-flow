from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json
from decimal import Decimal
from ..models import StageLimit, Transaction

@login_required
def concrete_analytics(request):
    """
    Звіт по Бетону.
    Додано: Графіки, фікс прогрес-барів, характеристики.
    Виправлено: Обробка відсутнього поля construct_type.
    """
    # 1. Шукаємо ліміти по бетону (або за назвою, або за категорією)
    limits = StageLimit.objects.filter(
        material__name__icontains='бетон'
    ).select_related('stage', 'material', 'stage__warehouse', 'material__category').order_by('stage__name')
    
    report_data = []
    total_plan = Decimal("0.000")
    total_fact = Decimal("0.000")
    
    # Дані для графіка
    chart_labels = []
    chart_plan_data = []
    chart_fact_data = []
    
    for limit in limits:
        # Рахуємо фактичне відпрацювання
        fact_spent = Transaction.objects.filter(
            warehouse=limit.stage.warehouse,
            material=limit.material,
            transaction_type='OUT',
            stage=limit.stage
        ).aggregate(s=Sum('quantity'))['s'] or Decimal("0.000")
        
        plan = limit.planned_quantity
        fact = fact_spent
        diff = plan - fact
        
        status = 'ok'
        if diff < 0: 
            status = 'over'
        elif plan > 0 and diff < (plan * Decimal("0.1")): 
            status = 'warning'
        
        percent = (fact / plan * 100) if plan > 0 else 0
        
        # ВИПРАВЛЕННЯ: Безпечне отримання типу конструкції
        # Оскільки поля construct_type немає в StageLimit, беремо його з категорії матеріалу або ставимо дефолт
        c_type_name = "Загальнобуд."
        
        # Спроба отримати з пов'язаних полів, якщо вони існують
        if hasattr(limit, 'construct_type'):
            c_type_name = limit.construct_type
        elif limit.material.category:
            c_type_name = limit.material.category.name
        
        report_data.append({
            'warehouse': limit.stage.warehouse.name,
            'stage': limit.stage.name,
            'material': limit.material.name,
            'characteristics': limit.material.characteristics,
            'unit': limit.material.unit,
            'plan': plan,
            'fact': fact,
            'diff': diff,
            'percent': int(percent),
            'status': status,
            'type': c_type_name
        })
        
        total_plan += plan
        total_fact += fact

        # Data for chart
        chart_labels.append(f"{limit.stage.name} ({limit.material.name})")
        # Chart.js потребує float/int
        chart_plan_data.append(float(plan))
        chart_fact_data.append(float(fact))

    # EXPORT TO EXCEL
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Звіт по бетону"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center')
        
        headers = ['Об\'єкт', 'Етап', 'Тип', 'Матеріал', 'Характеристики', 'План (м3)', 'Факт (м3)', 'Різниця', 'Статус']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row in report_data:
            st = "Норма"
            if row['status'] == 'over': st = "ПЕРЕВИТРАТА"
            elif row['status'] == 'warning': st = "Увага"

            ws.append([
                row['warehouse'], row['stage'], row['type'], 
                row['material'], row['characteristics'],
                row['plan'], row['fact'], row['diff'], st
            ])
            
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Concrete_Report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    return render(request, 'warehouse/concrete_report.html', {
        'report_data': report_data,
        'total_plan': total_plan,
        'total_fact': total_fact,
        'total_diff': total_plan - total_fact, # Економія (позитивне) або перевитрата (негативне)
        # JSON для графіка
        'chart_labels': json.dumps(chart_labels),
        'chart_plan_data': json.dumps(chart_plan_data),
        'chart_fact_data': json.dumps(chart_fact_data),
    })