from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from django.db.models import Q
import json
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from decimal import Decimal

logger = logging.getLogger('warehouse')

from ..models import Order, OrderItem, Warehouse, Material, Supplier, AuditLog, Transaction
from ..forms import OrderForm, OrderItemFormSet
from ..services import inventory
from ..services.inventory import InsufficientStockError
from .utils import log_audit, check_access, capture_order_snapshot, compute_order_diff
from ..decorators import rate_limit, staff_required

# ==============================================================================
# СПИСОК ЗАЯВОК (ORDER LIST)
# ==============================================================================

@login_required
def order_list(request):
    """
    Загальний список заявок з фільтрацією.
    Відображає всі заявки, відсортовані від найновіших.
    """
    # Оптимізація: завантажуємо пов'язані об'єкти (склад, автор) та товари (items + матеріали)
    orders = Order.objects.select_related('warehouse', 'created_by', 'source_warehouse') \
                          .prefetch_related('items__material') \
                          .order_by('-created_at')

    # Фільтрація по статусу
    status = request.GET.get('status')
    if status:
        orders = orders.filter(status=status)

    # Фільтрація по датах
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)

    # EXPORT TO EXCEL
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Журнал заявок"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = ['ID', 'Дата', 'Тип', 'Об\'єкт', 'Статус', 'Пріоритет', 'Автор', 'Позицій', 'Сума (грн)']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data
        for order in orders:
            order_type = "Переміщення" if order.source_warehouse else "Закупівля"
            total_sum = sum(item.quantity * (item.material.current_avg_price or 0) for item in order.items.all())
            ws.append([
                order.id,
                order.created_at.strftime('%d.%m.%Y'),
                order_type,
                order.warehouse.name,
                order.get_status_display(),
                order.get_priority_display(),
                order.created_by.get_full_name() if order.created_by else "—",
                order.items.count(),
                float(total_sum)
            ])

        # Autosize columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Orders_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    return render(request, 'warehouse/order_list.html', {
        'orders': orders,
        'f_status': status,
        'f_date_from': date_from,
        'f_date_to': date_to
    })


# ==============================================================================
# СТВОРЕННЯ ТА РЕДАГУВАННЯ (CREATE / EDIT)
# ==============================================================================

@login_required
def create_order(request):
    """
    Створення нової заявки.
    """
    if request.method == 'POST':
        form = OrderForm(request.POST, request.FILES, user=request.user)
        formset = OrderItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    order = form.save(commit=False)
                    order.created_by = request.user
                    order.save()

                    formset.instance = order
                    formset.save()

                    log_audit(request, 'CREATE', order, new_val="Створено заявку")
                    messages.success(request, f"Заявку #{order.id} успішно створено!")

                    if request.user.is_staff:
                        return redirect('manager_order_detail', pk=order.id)
                    else:
                        return redirect('foreman_order_detail', pk=order.id)
            except (ValidationError, ValueError) as e:
                messages.error(request, f"Помилка валідації: {e}")
            except Exception as e:
                logger.exception(f"Order creation failed for user {request.user.id}")
                messages.error(request, "Помилка при створенні заявки. Спробуйте ще раз.")
    else:
        form = OrderForm(user=request.user)
        formset = OrderItemFormSet()

    return render(request, 'warehouse/create_order.html', {
        'form': form,
        'formset': formset,
        'edit_mode': False
    })

@login_required
def edit_order(request, pk):
    """
    Редагування існуючої заявки.
    """
    order = get_object_or_404(Order, pk=pk)
    
    # Перевірка прав: редагувати може автор або менеджер
    if not request.user.is_staff and order.created_by != request.user:
        return HttpResponse("⛔ Немає доступу до редагування цієї заявки", status=403)
        
    # Заборона редагування завершених заявок (опціонально)
    if order.status in ['completed', 'rejected']:
        messages.warning(request, "Цю заявку вже не можна редагувати, оскільки вона закрита.")
        # Редірект на перегляд
        if request.user.is_staff:
            return redirect('manager_order_detail', pk=order.id)
        else:
            return redirect('foreman_order_detail', pk=order.id)

    if request.method == 'POST':
        form = OrderForm(request.POST, request.FILES, instance=order, user=request.user)
        formset = OrderItemFormSet(request.POST, instance=order)

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    old_snapshot = capture_order_snapshot(order)
                    form.save()
                    formset.save()
                    changed = compute_order_diff(old_snapshot, order)

                    log_audit(request, 'UPDATE', order,
                              new_val="Редаговано заявку",
                              changed_fields=changed or None)
                    messages.success(request, f"Заявку #{order.id} успішно оновлено!")

                    if request.user.is_staff:
                        return redirect('manager_order_detail', pk=order.id)
                    else:
                        return redirect('foreman_order_detail', pk=order.id)
            except (ValidationError, ValueError) as e:
                messages.error(request, f"Помилка валідації: {e}")
            except Exception as e:
                logger.exception(f"Order edit failed for order {order.id}, user {request.user.id}")
                messages.error(request, "Помилка при збереженні. Спробуйте ще раз.")
    else:
        form = OrderForm(instance=order, user=request.user)
        formset = OrderItemFormSet(instance=order)

    return render(request, 'warehouse/create_order.html', {
        'form': form,
        'formset': formset,
        'edit_mode': True
    })

@login_required
def delete_order(request, pk):
    """
    М'яке видалення заявки (переміщення до кошика).
    """
    order = get_object_or_404(Order, pk=pk)

    if not request.user.is_staff and order.created_by != request.user:
        return HttpResponse("⛔ Немає доступу", status=403)

    if order.status not in ('new', 'rejected'):
        messages.error(request, "До кошика можна перемістити лише нову або відхилену заявку.")
    else:
        log_audit(request, 'DELETE', order, old_val=f"Order #{order.id} moved to trash")
        order.delete()  # soft delete
        messages.success(request, f"Заявку #{order.id} переміщено до кошика.")

    if request.user.is_staff:
        return redirect('manager_dashboard')
    return redirect('index')


# ==============================================================================
# ЛОГІСТИКА
# ==============================================================================

@staff_required
def logistics_monitor(request):
    """
    Монітор логіста: заявки в статусі 'purchasing' (треба везти) та 'transit' (їдуть).
    """
    purchasing_orders = Order.objects.filter(status='purchasing').order_by('expected_date')
    transit_orders = Order.objects.filter(status='transit').order_by('expected_date')
    
    return render(request, 'warehouse/logistics.html', {
        'purchasing_orders': purchasing_orders,
        'transit_orders': transit_orders
    })

@staff_required
def mark_order_shipped(request, pk):
    """
    Логіст позначає, що товар виїхав (status -> transit).
    """
    order = get_object_or_404(Order, pk=pk)
    
    if request.method == 'POST':
        # Тут можна обробити дані водія/авто з форми
        driver_phone = request.POST.get('driver_phone')
        vehicle_number = request.POST.get('vehicle_number')
        
        # Можна зберегти це в примітку
        note_add = f"\n[Логістика] Водій: {driver_phone}, Авто: {vehicle_number}"
        order.note += note_add
        
        old_status = order.get_status_display()
        order.status = 'transit'
        order._actor = request.user
        order.save()

        log_audit(request, 'ORDER_STATUS', order,
                  changed_fields={'status': {
                      'old': old_status, 'new': order.get_status_display(), 'label': 'Статус',
                  }})
        messages.success(request, f"Заявку #{order.id} відправлено (Transit).")
        
    return redirect('logistics_monitor')

@login_required
def confirm_receipt(request, pk):
    """
    Підтвердження отримання (на складі).
    GET  — показує форму прийому (confirm_receipt.html).
    POST — викликає сервіс inventory.process_order_receipt і редіректить.
    """
    order = get_object_or_404(Order, pk=pk)

    # Перевірка доступу до складу
    if not check_access(request.user, order.warehouse):
        return HttpResponse("Немає доступу до складу цієї заявки", status=403)

    # Прийом можливий лише коли товар в дорозі або ще не відправлений (purchasing)
    if order.status not in ('transit', 'purchasing', 'approved'):
        messages.warning(request, f"Заявку #{order.id} не можна прийняти — статус: {order.get_status_display()}.")
        if request.user.is_staff:
            return redirect('manager_order_detail', pk=order.id)
        return redirect('foreman_order_detail', pk=order.id)

    if request.method == 'POST':
        # Перевіряємо, що item_id належить саме цій заявці (захист від IDOR)
        valid_item_ids = set(str(i) for i in order.items.values_list('id', flat=True))
        items_data = {}
        for key, value in request.POST.items():
            if key.startswith('item_qty_'):
                item_id = key.split('_')[-1]
                if item_id in valid_item_ids:
                    items_data[item_id] = value

        proof_photo = request.FILES.get('proof_photo')
        comment = request.POST.get('comment', '')

        try:
            inventory.process_order_receipt(order, items_data, request.user, proof_photo, comment)
            log_audit(request, 'ORDER_RECEIVED', order, new_val="Items added to stock")
            messages.success(request, f"Заявку #{order.id} успішно прийнято на склад!")
            if request.user.is_staff:
                return redirect('manager_order_detail', pk=order.id)
            return redirect('foreman_order_detail', pk=order.id)

        except InsufficientStockError as e:
            messages.error(request, f"Недостатньо товару на складі: {e.material.name}")
        except (ValidationError, ValueError) as e:
            messages.error(request, f"Помилка валідації: {e}")
        except Exception as e:
            logger.exception(f"Order receipt failed for order {order.id}, user {request.user.id}")
            messages.error(request, "Помилка при прийомі товару. Спробуйте ще раз.")

    # GET або POST з помилкою — показуємо форму
    return render(request, 'warehouse/confirm_receipt.html', {'order': order})


# ==============================================================================
# AJAX & UTILS
# ==============================================================================

@login_required
@rate_limit(requests_per_minute=30, key_prefix='ajax_order_dup')
def check_order_duplicates(request):
    """
    AJAX: Перевіряє, чи не створювали схожу заявку на цей склад недавно.
    """
    wh_id = request.GET.get('warehouse')
    if not wh_id: return JsonResponse({'exists': False})

    # Шукаємо заявки за останні 3 дні
    three_days_ago = timezone.now() - timezone.timedelta(days=3)
    recent_orders = Order.objects.filter(
        warehouse_id=wh_id,
        created_at__gte=three_days_ago
    ).exclude(status__in=['completed', 'rejected', 'draft']).order_by('-created_at')
    
    if recent_orders.exists():
        data = []
        for o in recent_orders:
            items_str = ", ".join([i.material.name for i in o.items.all()[:3]])
            data.append({
                'id': o.id,
                'date': o.created_at.strftime("%d.%m %H:%M"),
                'items': items_str
            })
        return JsonResponse({'exists': True, 'orders': data})
        
    return JsonResponse({'exists': False})

@login_required
def print_order_pdf(request, pk):
    """
    Генерація сторінки для друку заявки.
    """
    order = get_object_or_404(Order, pk=pk)
    # Перевірка доступу
    if not request.user.is_staff and not check_access(request.user, order.warehouse):
        return HttpResponse("⛔ Немає доступу", status=403)
        
    return render(request, 'warehouse/print_order.html', {'order': order})


@login_required
def print_order_qr_labels(request, pk):
    """
    Сторінка друку QR-етикеток для матеріалів заявки.
    QR-код кодує 'MAT:{material.pk}' для сканування при прийомі товару.
    """
    order = get_object_or_404(Order, pk=pk)
    if not request.user.is_staff and not check_access(request.user, order.warehouse):
        return HttpResponse("⛔ Немає доступу", status=403)
    items = order.items.select_related('material').all()
    return render(request, 'warehouse/print_order_qr_labels.html', {
        'order': order,
        'items': items,
    })


# ==============================================================================
# КОШИК (SOFT-DELETE TRASH)
# ==============================================================================

@staff_required
def trash_view(request):
    """
    Кошик: відображає м'яко видалені заявки та транзакції.
    Доступний тільки для персоналу (is_staff).
    """
    deleted_orders = (
        Order.all_objects
        .filter(is_deleted=True)
        .select_related('created_by', 'warehouse')
        .order_by('-deleted_at')
    )
    deleted_transactions = (
        Transaction.all_objects
        .filter(is_deleted=True)
        .select_related('material', 'warehouse', 'created_by')
        .order_by('-deleted_at')
    )
    return render(request, 'warehouse/trash.html', {
        'deleted_orders': deleted_orders,
        'deleted_transactions': deleted_transactions,
        'page_title': 'Кошик',
    })


@staff_required
def restore_order(request, pk):
    """Відновлення заявки з кошика."""
    order = Order.all_objects.filter(pk=pk, is_deleted=True).first()
    if not order:
        messages.error(request, "Заявку не знайдено в кошику.")
        return redirect('trash')
    order.restore()
    log_audit(request, 'UPDATE', order, new_val=f"Order #{order.id} restored from trash")
    messages.success(request, f"Заявку #{order.id} відновлено.")
    return redirect('trash')


@staff_required
def delete_order_permanent(request, pk):
    """Остаточне (фізичне) видалення заявки з кошика."""
    order = Order.all_objects.filter(pk=pk, is_deleted=True).first()
    if not order:
        messages.error(request, "Заявку не знайдено в кошику.")
        return redirect('trash')
    if request.method == 'POST':
        order_id = order.id
        log_audit(request, 'DELETE', order, old_val=f"Order #{order.id} permanently deleted")
        order.hard_delete()
        messages.success(request, f"Заявку #{order_id} остаточно видалено.")
    return redirect('trash')


@staff_required
def delete_transaction(request, pk):
    """М'яке видалення транзакції (переміщення до кошика)."""
    txn = get_object_or_404(Transaction, pk=pk)
    log_audit(request, 'DELETE', txn, old_val=f"Transaction #{txn.id} moved to trash")
    txn.delete()  # soft delete
    messages.success(request, f"Транзакцію #{txn.id} переміщено до кошика.")
    next_url = request.POST.get('next') or request.GET.get('next') or 'index'
    return redirect(next_url)


@staff_required
def restore_transaction(request, pk):
    """Відновлення транзакції з кошика."""
    txn = Transaction.all_objects.filter(pk=pk, is_deleted=True).first()
    if not txn:
        messages.error(request, "Транзакцію не знайдено в кошику.")
        return redirect('trash')
    txn.restore()
    log_audit(request, 'UPDATE', txn, new_val=f"Transaction #{txn.id} restored from trash")
    messages.success(request, f"Транзакцію #{txn.id} відновлено.")
    return redirect('trash')


@staff_required
def delete_transaction_permanent(request, pk):
    """Остаточне (фізичне) видалення транзакції з кошика."""
    txn = Transaction.all_objects.filter(pk=pk, is_deleted=True).first()
    if not txn:
        messages.error(request, "Транзакцію не знайдено в кошику.")
        return redirect('trash')
    if request.method == 'POST':
        txn_id = txn.id
        log_audit(request, 'DELETE', txn, old_val=f"Transaction #{txn.id} permanently deleted")
        txn.hard_delete()
        messages.success(request, f"Транзакцію #{txn_id} остаточно видалено.")
    return redirect('trash')