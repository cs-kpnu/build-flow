from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json
from decimal import Decimal
from ..models import StageLimit, Transaction

@login_required
def rebar_analytics(request):
    """
    Звіт по Арматурі (Металу).
    Додано: Графіки, фікс прогрес-барів.
    Виправлено: AttributeError 'StageLimit' object has no attribute 'construct_type'
    """
    # 1. Шукаємо ліміти, де матеріал містить "арматура" або "armatura"
    # Оптимізація: додано material__category
    limits = StageLimit.objects.filter(
        material__name__icontains='арматура'
    ).select_related('stage', 'material', 'stage__warehouse', 'material__category').order_by('stage__name')
    
    report_data = []
    total_plan = Decimal("0.000")
    total_fact = Decimal("0.000")
    
    # Масиви для графіка Chart.js
    chart_labels = []
    chart_plan_data = []
    chart_fact_data = []
    
    for limit in limits:
        # Рахуємо фактичне відпрацювання
        fact_spent = Transaction.objects.filter(
            warehouse=limit.stage.warehouse,
            material=limit.material,
            transaction_type='OUT',
            stage=limit.stage
        ).aggregate(s=Sum('quantity'))['s'] or Decimal("0.000")
        
        plan = limit.planned_quantity
        fact = fact_spent
        # diff = План - Факт. Якщо > 0 — економія, < 0 — перевитрата.
        diff = plan - fact
        
        status = 'ok'
        if diff < 0: 
            status = 'over'
        # Попередження, якщо залишилось менше 10% запасу (але ще не нуль)
        elif plan > 0 and diff < (plan * Decimal("0.1")): 
            status = 'warning'
        
        percent = (fact / plan * 100) if plan > 0 else 0
        
        # ВИПРАВЛЕННЯ: Безпечне отримання типу конструкції
        # Оскільки поля construct_type немає, беремо назву категорії матеріалу або дефолт
        c_type_name = "Без категорії"
        if limit.material.category:
            c_type_name = limit.material.category.name
            
        report_data.append({
            'warehouse': limit.stage.warehouse.name,
            'stage': limit.stage.name,
            'type': c_type_name, 
            'material': limit.material.name,
            'characteristics': limit.material.characteristics,
            'unit': limit.material.unit,
            'plan': plan,
            'fact': fact,
            'diff': diff,
            'percent': int(percent),
            'status': status
        })
        
        total_plan += plan
        total_fact += fact
        
        # Data for chart
        chart_labels.append(f"{limit.stage.name} ({limit.material.name})")
        chart_plan_data.append(float(plan))
        chart_fact_data.append(float(fact))

    # EXPORT TO EXCEL
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Звіт по арматурі"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center')
        
        headers = ['Об\'єкт', 'Етап', 'Тип', 'Матеріал', 'Характеристики', 'Од.', 'План', 'Факт', 'Різниця', 'Статус']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row in report_data:
            status_text = "Норма"
            if row['status'] == 'over': status_text = "ПЕРЕВИТРАТА"
            elif row['status'] == 'warning': status_text = "Увага"

            ws.append([
                row['warehouse'],
                row['stage'],
                row['type'],
                row['material'],
                row['characteristics'],
                row['unit'],
                row['plan'],
                row['fact'],
                row['diff'],
                status_text
            ])
            
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Rebar_Report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    return render(request, 'warehouse/rebar_report.html', {
        'report_data': report_data,
        'total_plan': total_plan,
        'total_fact': total_fact,
        # total_diff: якщо > 0 це економія (план > факт), якщо < 0 це перевитрата
        'total_diff': total_plan - total_fact,
        # JSON для графіка
        'chart_labels': json.dumps(chart_labels),
        'chart_plan_data': json.dumps(chart_plan_data),
        'chart_fact_data': json.dumps(chart_fact_data),
    })