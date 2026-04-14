from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.views.decorators.http import require_POST, require_GET
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import io
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger('warehouse')

from ..models import Order, UserProfile, Warehouse, ConstructionStage, Material, Transaction, Category
from ..forms import UserUpdateForm, ProfileUpdateForm
from .utils import get_user_warehouses, get_warehouse_balance, check_access, log_audit
from ..decorators import rate_limit, staff_required

# ==============================================================================
# ГОЛОВНА СТОРІНКА
# ==============================================================================

@login_required
def index(request):
    """
    Точка входу в систему.
    Визначає роль користувача і відображає відповідний дашборд.
    """
    user_warehouses = get_user_warehouses(request.user)
    
    # Базовий контекст
    context = {
        'warehouses': user_warehouses,
        'now': timezone.now()
    }
    
    # --- ЛОГІКА ДЛЯ ПРОРАБА (Non-Staff) ---
    if not request.user.is_staff:
        context['role'] = 'foreman'
        
        # Перевіряємо, чи вибрано активний склад в сесії
        active_wh_id = request.session.get('active_warehouse_id')
        active_wh = None
        
        if active_wh_id:
            try:
                active_wh = user_warehouses.get(pk=active_wh_id)
            except Warehouse.DoesNotExist:
                pass
        
        # Якщо склад не вибрано або він недоступний, беремо перший доступний
        if not active_wh and user_warehouses.exists():
            active_wh = user_warehouses.first()
            request.session['active_warehouse_id'] = active_wh.id
            
        context['active_warehouse'] = active_wh
        
        # Розрахунок метрик для дашборда (items_count, total_value)
        # Логіка ідентична foreman_storage_view
        items_count = 0
        total_value = Decimal("0.00")
        
        if active_wh:
            # Отримуємо баланс: {Material: Decimal(qty)}
            balance_map = get_warehouse_balance(active_wh)
            
            for material, qty in balance_map.items():
                if qty > 0: # Враховуємо тільки позитивні залишки
                    items_count += 1
                    
                    # Ціна та сума
                    avg_price = material.current_avg_price or Decimal("0.00")
                    sum_val = (qty * avg_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    total_value += sum_val
        
        context['items_count'] = items_count
        context['total_value'] = total_value
        
        # Останні активні заявки для цього складу
        if active_wh:
            context['my_orders'] = Order.objects.filter(
                warehouse=active_wh
            ).exclude(status__in=['completed', 'rejected']).order_by('-updated_at')[:5]
            
        return render(request, 'warehouse/index.html', context)

    # --- ЛОГІКА ДЛЯ МЕНЕДЖЕРА/АДМІНА (Staff) ---
    else:
        context['role'] = 'manager'
        # Менеджери мають свій дашборд
        return redirect('manager_dashboard')


# ==============================================================================
# ПРОФІЛЬ КОРИСТУВАЧА
# ==============================================================================

@login_required
def profile_view(request):
    """
    Сторінка профілю користувача.
    """
    if request.method == 'POST':
        u_form = UserUpdateForm(request.POST, instance=request.user)
        p_form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user.profile)
        
        if u_form.is_valid() and p_form.is_valid():
            u_form.save()
            p_form.save()
            messages.success(request, 'Ваш профіль успішно оновлено!')
            return redirect('profile')
    else:
        u_form = UserUpdateForm(instance=request.user)
        # Переконаємось, що профіль існує (сигнали повинні були створити його, але для надійності)
        if not hasattr(request.user, 'profile'):
            UserProfile.objects.create(user=request.user)
            
        p_form = ProfileUpdateForm(instance=request.user.profile)

    context = {
        'u_form': u_form,
        'p_form': p_form
    }
    return render(request, 'warehouse/profile.html', context)


@login_required
def change_password_view(request):
    """
    Зміна пароля.
    """
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Щоб не розлогінило
            messages.success(request, 'Пароль успішно змінено!')
            return render(request, 'warehouse/password_change_done.html')
        else:
            messages.error(request, 'Виправте помилки нижче.')
    else:
        form = PasswordChangeForm(request.user)
        
    return render(request, 'warehouse/password_change_form.html', {'form': form})


# ==============================================================================
# УПРАВЛІННЯ СЕСІЄЮ
# ==============================================================================

@login_required
def switch_active_warehouse(request, pk):
    """
    Швидке перемикання активного складу.
    """
    if check_access(request.user, pk):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        request.session['active_warehouse_id'] = pk
        messages.success(request, f"🏢 Активний об'єкт змінено на: {warehouse.name}")
    else:
        messages.error(request, "⛔ У вас немає доступу до цього складу.")
        
    referer = request.META.get('HTTP_REFERER', '')
    if referer and url_has_allowed_host_and_scheme(referer, allowed_hosts={request.get_host()}):
        return redirect(referer)
    return redirect('index')


# ==============================================================================
# КАТАЛОГ МАТЕРІАЛІВ
# ==============================================================================

@login_required
def material_list(request):
    """
    Список матеріалів з пошуком та пагінацією.
    """
    query = request.GET.get('q', '')
    materials_list = Material.objects.all().select_related('category').order_by('name')
    
    if query:
        materials_list = materials_list.filter(
            Q(name__icontains=query) | 
            Q(article__icontains=query) |
            Q(characteristics__icontains=query)
        )
        
    paginator = Paginator(materials_list, 20) # 20 матеріалів на сторінку
    page_number = request.GET.get('page')
    materials = paginator.get_page(page_number)
    
    return render(request, 'warehouse/material_list.html', {
        'materials': materials,
        'page_title': 'Довідник матеріалів'
    })


@login_required
def material_detail(request, pk):
    """
    Детальна сторінка матеріалу.
    Показує загальні залишки по всіх складах та історію руху.
    """
    material = get_object_or_404(Material, pk=pk)
    
    # 1. Рахуємо залишки по складах (тільки доступні користувачу, або всі для менеджера)
    warehouses = get_user_warehouses(request.user)
    stock_distribution = []
    total_quantity = 0

    # Батч-агрегація замість N+1 запитів у циклі
    agg_rows = Transaction.objects.filter(
        material=material,
        warehouse__in=warehouses,
        transaction_type__in=['IN', 'OUT', 'LOSS']
    ).values('warehouse_id').annotate(
        in_qty=Sum('quantity', filter=Q(transaction_type='IN')),
        out_qty=Sum('quantity', filter=Q(transaction_type__in=['OUT', 'LOSS']))
    )
    wh_map = {wh.id: wh for wh in warehouses}
    for row in agg_rows:
        qty = (row['in_qty'] or 0) - (row['out_qty'] or 0)
        if qty != 0:
            wh = wh_map.get(row['warehouse_id'])
            stock_distribution.append({
                'warehouse': wh.name if wh else '—',
                'quantity': round(qty, 2)
            })
            total_quantity += qty
            
    # 2. Оціночна вартість
    total_value = float(total_quantity) * float(material.current_avg_price or 0)
    
    # 3. Історія операцій (останні 50)
    # Фільтруємо транзакції тільки по доступних складах
    transactions = Transaction.objects.filter(
        material=material,
        warehouse__in=warehouses
    ).select_related('warehouse', 'created_by').order_by('-created_at')[:50]

    return render(request, 'warehouse/material_detail.html', {
        'material': material,
        'stock_distribution': stock_distribution,
        'total_quantity': round(total_quantity, 2),
        'total_value': round(total_value, 2),
        'transactions': transactions
    })


# ==============================================================================
# AJAX API
# ==============================================================================

@login_required
@require_GET
@rate_limit(requests_per_minute=60, key_prefix='ajax_stages')
def load_stages(request):
    """
    API: Повертає список етапів будівництва для вибраного складу.
    URL: /ajax/load-stages/?warehouse_id=123
    """
    warehouse_id = request.GET.get('warehouse_id')
    stages = []
    
    if warehouse_id:
        try:
            wh_id_int = int(warehouse_id)
            if check_access(request.user, wh_id_int):
                qs = ConstructionStage.objects.filter(warehouse_id=wh_id_int).order_by('name')
                stages = list(qs.values('id', 'name'))
        except (ValueError, TypeError) as e:
            logger.debug(f"load_stages: некоректний warehouse_id {warehouse_id!r}: {e}")
            
    return JsonResponse(stages, safe=False)

@login_required
@require_GET
@rate_limit(requests_per_minute=120, key_prefix='ajax_mat_general')
def ajax_materials(request):
    """
    API: Пошук матеріалів для TomSelect (Autocomplete).
    URL: /ajax/materials/?q=query
    """
    query = request.GET.get('q', '')
    materials = Material.objects.all().order_by('name')
    
    if query:
        materials = materials.filter(name__icontains=query)
    
    # Повертаємо топ-50 результатів
    results = list(materials.values('id', 'name')[:50])
    return JsonResponse(results, safe=False)


# ==============================================================================
# МАСОВИЙ ІМПОРТ МАТЕРІАЛІВ З EXCEL
# ==============================================================================

# Очікувані заголовки колонок (нечутливі до регістру та пробілів)
_IMPORT_COLUMNS = ['назва', 'артикул', 'одиниця', 'категорія', 'характеристики', 'мін. залишок', 'середня ціна']

_HEADER_STYLE = {
    'font': Font(bold=True, color='FFFFFF'),
    'fill': PatternFill(fill_type='solid', fgColor='1F3864'),
    'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
}


def _safe_decimal(value, default=Decimal('0')):
    """Безпечно конвертує значення в Decimal ≥ 0."""
    if value is None or str(value).strip() == '':
        return default
    try:
        d = Decimal(str(value).replace(',', '.').strip())
        return max(d, Decimal('0'))
    except InvalidOperation:
        return None  # сигнал про помилку


def _normalize_header(cell_value):
    return str(cell_value or '').strip().lower()


@staff_required
def import_materials_template(request):
    """Завантаження порожнього Excel-шаблону для імпорту матеріалів."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Матеріали'

    headers = ['Назва *', 'Артикул', 'Одиниця (шт за замовч.)', 'Категорія', 'Характеристики', 'Мін. залишок', 'Середня ціна (грн)']
    col_widths = [40, 18, 22, 20, 35, 16, 20]

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_STYLE['font']
        cell.fill = _HEADER_STYLE['fill']
        cell.alignment = _HEADER_STYLE['alignment']
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # Приклад рядка
    example = ['Цемент М400', 'CEM-400', 'кг', 'Будматеріали', 'Портландцемент, мішок 25 кг', '500', '4.50']
    example_fill = PatternFill(fill_type='solid', fgColor='EBF1DE')
    for col_idx, val in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.fill = example_fill
        cell.border = border

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="materials_import_template.xlsx"'
    return response


@staff_required
def import_materials(request):
    """
    GET  — форма завантаження Excel.
    POST — обробка файлу: upsert матеріалів по артикулу або назві.

    Логіка upsert:
    - Якщо артикул заповнений і вже існує → UPDATE
    - Якщо артикул заповнений і новий      → CREATE
    - Якщо артикул порожній                → пошук по точному імені → UPDATE або CREATE
    """
    if request.method != 'POST':
        return render(request, 'warehouse/import_materials.html')

    uploaded = request.FILES.get('excel_file')
    if not uploaded:
        messages.error(request, 'Будь ласка, оберіть файл Excel.')
        return render(request, 'warehouse/import_materials.html')

    if not uploaded.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Допускаються тільки файли .xlsx або .xls.')
        return render(request, 'warehouse/import_materials.html')

    try:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
    except Exception:
        messages.error(request, 'Не вдалося відкрити файл. Переконайтесь, що це коректний Excel.')
        return render(request, 'warehouse/import_materials.html')

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, 'Файл порожній.')
        return render(request, 'warehouse/import_materials.html')

    # Визначаємо заголовки (перший рядок)
    raw_headers = [_normalize_header(h) for h in rows[0]]

    # Маппінг колонок: canonical_name → індекс
    col_map = {}
    for idx, h in enumerate(raw_headers):
        h_clean = h.replace('*', '').replace('(шт за замовч.)', '').replace('(грн)', '').strip()
        if 'назва' in h_clean:
            col_map['name'] = idx
        elif 'артикул' in h_clean:
            col_map['article'] = idx
        elif 'одиниц' in h_clean:
            col_map['unit'] = idx
        elif 'категорі' in h_clean:
            col_map['category'] = idx
        elif 'характерист' in h_clean:
            col_map['characteristics'] = idx
        elif 'мін' in h_clean:
            col_map['min_limit'] = idx
        elif 'середн' in h_clean or 'ціна' in h_clean:
            col_map['current_avg_price'] = idx

    if 'name' not in col_map:
        messages.error(request, 'Колонка "Назва" не знайдена. Перевірте заголовки файлу.')
        return render(request, 'warehouse/import_materials.html')

    # Кеш категорій (get_or_create по імені)
    category_cache = {}

    def get_category(cat_name):
        if not cat_name:
            return None
        cat_name = str(cat_name).strip()
        if not cat_name:
            return None
        if cat_name not in category_cache:
            cat, _ = Category.objects.get_or_create(name=cat_name)
            category_cache[cat_name] = cat
        return category_cache[cat_name]

    # Обробка рядків
    created = updated = skipped = 0
    row_errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        def get_col(key):
            idx = col_map.get(key)
            if idx is None:
                return None
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val is not None else ''

        name = get_col('name')
        if not name:
            skipped += 1
            continue  # порожній рядок

        article = get_col('article') or None
        unit = get_col('unit') or 'шт'
        category = get_category(get_col('category'))
        characteristics = get_col('characteristics') or ''

        min_limit = _safe_decimal(get_col('min_limit'), Decimal('0'))
        avg_price = _safe_decimal(get_col('current_avg_price'), Decimal('0'))

        if min_limit is None:
            row_errors.append(f'Рядок {row_num}: некоректний "Мін. залишок" — пропущено.')
            skipped += 1
            continue
        if avg_price is None:
            row_errors.append(f'Рядок {row_num}: некоректна "Середня ціна" — пропущено.')
            skipped += 1
            continue

        # Знаходимо або створюємо матеріал
        material = None
        is_new = False

        if article:
            material = Material.objects.filter(article=article).first()
        if material is None:
            material = Material.objects.filter(name=name).first()
        if material is None:
            material = Material()
            is_new = True

        material.name = name
        if article:
            material.article = article
        material.unit = unit
        material.characteristics = characteristics
        material.category = category
        material.min_limit = min_limit.quantize(Decimal('0.001'))
        # Середню ціну оновлюємо лише якщо явно задана (не 0) або новий матеріал
        if avg_price > 0 or is_new:
            material.current_avg_price = avg_price.quantize(Decimal('0.01'))

        try:
            material.save()
            if is_new:
                created += 1
            else:
                updated += 1
        except Exception as e:
            row_errors.append(f'Рядок {row_num} ({name}): помилка збереження — {e}')
            skipped += 1

    log_audit(request, 'CREATE', new_val=f'Excel import: +{created} created, ~{updated} updated, {skipped} skipped')

    return render(request, 'warehouse/import_materials.html', {
        'result': {
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': row_errors,
            'total': created + updated + skipped,
        }
    })